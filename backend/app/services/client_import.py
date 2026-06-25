"""Client-import: parse a multi-tab XLSX (or CSV bundle) describing clients,
their projects, tasks, and who is assigned to what, then preview + commit.

File shape (XLSX with up to 4 sheets; sheet names matched case-insensitively):
  Clients      : name | company | type | status | contact_name | contact_email
                 | contact_phone | since
  Projects     : client | name | code | billable_rate | budget | currency
                 | status | start_date | end_date | managers   (managers =
                 semicolon-separated emails of existing MANAGER/ADMIN users)
  Tasks        : client | project | name | priority | status | description
  Assignments  : client | project | task | user_email      (task blank = grant
                 the whole project to the user)

A single CSV is treated as the Projects sheet for the simplest case (one client
implied by the `client` column).

Everything is resolved/created in dependency order on commit: clients (deduped
by name) -> projects -> tasks -> assignments. Users are matched to EXISTING
accounts by email; unknown emails are reported, never invented.
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.client import Client
from app.models.project import Project
from app.models.task import Task
from app.models.user import User


# ── Canonical sheet + column definitions ────────────────────────────────────
SHEETS = {
    "clients": ["name", "company", "type", "status", "contact_name",
                "contact_email", "contact_phone", "since"],
    "projects": ["client", "name", "code", "billable_rate", "budget",
                 "currency", "status", "start_date", "end_date", "managers"],
    "tasks": ["client", "project", "name", "priority", "status", "description"],
    "assignments": ["client", "project", "task", "user_email"],
}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _key(s: str) -> str:
    return s.strip().lower().replace(" ", "_")


def _rows_to_dicts(headers: list[str], rows: list[list[str]], cols: list[str]) -> list[dict]:
    """Map a sheet's rows to dicts keyed by our canonical column names, matching
    headers case/space-insensitively. Unknown headers are ignored; missing
    columns come back as ''. Blank rows are dropped."""
    idx = {}
    for i, h in enumerate(headers):
        idx[_key(_norm(h))] = i
    out: list[dict] = []
    for row in rows:
        if not any(_norm(c) for c in row):
            continue
        rec = {col: _norm(row[idx[col]]) if col in idx and idx[col] < len(row) else "" for col in cols}
        out.append(rec)
    return out


def parse_workbook(filename: str, content: bytes) -> dict[str, list[dict]]:
    """Parse the upload into {sheet_key: [row dicts]}. XLSX reads every matching
    sheet; CSV is treated as the Projects sheet."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        import csv
        import chardet
        enc = (chardet.detect(content) or {}).get("encoding") or "utf-8"
        text = content.decode(enc, errors="replace")
        reader = list(csv.reader(io.StringIO(text)))
        if not reader:
            return {k: [] for k in SHEETS}
        headers, rows = reader[0], reader[1:]
        return {
            "clients": [],
            "projects": _rows_to_dicts(headers, rows, SHEETS["projects"]),
            "tasks": [],
            "assignments": [],
        }
    if not (lower.endswith(".xlsx") or lower.endswith(".xls")):
        raise ValueError(f"Unsupported file type: {filename}. Upload a CSV or Excel (.xlsx) file.")

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    by_name = {_key(n): n for n in wb.sheetnames}
    result: dict[str, list[dict]] = {}
    for sheet_key, cols in SHEETS.items():
        actual = by_name.get(sheet_key)
        if actual is None:
            result[sheet_key] = []
            continue
        ws = wb[actual]
        raw = [[_norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
        if not raw:
            result[sheet_key] = []
            continue
        headers, rows = raw[0], raw[1:]
        result[sheet_key] = _rows_to_dicts(headers, rows, cols)
    wb.close()
    return result


# ── Resolution helpers (tenant-scoped) ──────────────────────────────────────
async def _client_by_name(db: AsyncSession, name: str, tenant_id: int) -> Optional[Client]:
    if not name:
        return None
    res = await db.execute(
        select(Client).where(Client.tenant_id == tenant_id, Client.name.ilike(name))
    )
    return res.scalars().first()


async def _user_by_email(db: AsyncSession, email: str, tenant_id: int) -> Optional[User]:
    if not email:
        return None
    res = await db.execute(
        select(User).where(User.tenant_id == tenant_id, User.email.ilike(email.strip()))
    )
    return res.scalars().first()


def _to_decimal(v: str) -> Optional[Decimal]:
    if not v:
        return None
    try:
        return Decimal(v.replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


# ── Preview: structure + per-row validation, NO writes ──────────────────────
async def build_preview(db: AsyncSession, data: dict[str, list[dict]], tenant_id: int) -> dict:
    """Validate the parsed data and return a summary + per-row notes. Does not
    write anything. Flags: unknown manager/assignee emails, missing required
    fields, references to a client/project not defined in the file or DB."""
    errors: list[str] = []
    warnings: list[str] = []

    clients = data.get("clients", [])
    projects = data.get("projects", [])
    tasks = data.get("tasks", [])
    assignments = data.get("assignments", [])

    # Client names known = those in the Clients sheet OR a Projects sheet's
    # `client` column (a project can imply a not-explicitly-listed client) OR
    # already in the DB.
    file_client_names = {c["name"].lower() for c in clients if c.get("name")}
    for p in projects:
        if p.get("client"):
            file_client_names.add(p["client"].lower())

    async def client_exists(name: str) -> bool:
        if not name:
            return False
        if name.lower() in file_client_names:
            return True
        return (await _client_by_name(db, name, tenant_id)) is not None

    # Per-row issues, keyed by sheet -> [{row, level, message}] where `row` is
    # the 0-based DATA row index (matches the preview table's row order, not the
    # spreadsheet line). level: "error" (blocks import) | "skip" (row dropped) |
    # "note" (imports with a change). Row numbers in messages use the 1-based
    # data row so they line up with the table.
    row_issues: dict[str, list[dict]] = {"clients": [], "projects": [], "tasks": [], "assignments": []}

    def flag(sheet: str, idx0: int, level: str, message: str) -> None:
        row_issues[sheet].append({"row": idx0, "level": level, "message": message})
        (errors if level == "error" else warnings).append(message)

    # Validate clients
    for i, c in enumerate(clients):
        if not c.get("name"):
            flag("clients", i, "error", f"Clients row {i + 1}: name is required.")

    # Validate projects + collect (client, project) keys
    project_keys: set[tuple[str, str]] = set()
    for i, p in enumerate(projects):
        if not p.get("name"):
            flag("projects", i, "error", f"Projects row {i + 1}: project name is required.")
        if not p.get("client"):
            flag("projects", i, "error", f"Projects row {i + 1}: client is required.")
        elif not await client_exists(p["client"]):
            flag("projects", i, "note", f"Projects row {i + 1}: client \"{p['client']}\" will be created.")
        if _to_decimal(p.get("billable_rate", "")) is None:
            flag("projects", i, "note", f"Projects row {i + 1}: billable rate missing/invalid — defaulting to 0.")
        # Manager emails must resolve to existing users.
        for em in [e.strip() for e in p.get("managers", "").replace(",", ";").split(";") if e.strip()]:
            if (await _user_by_email(db, em, tenant_id)) is None:
                flag("projects", i, "skip", f"Projects row {i + 1}: manager \"{em}\" isn't a user here — that PM is skipped.")
        if p.get("client") and p.get("name"):
            project_keys.add((p["client"].lower(), p["name"].lower()))

    # Validate tasks reference a known project
    for i, t in enumerate(tasks):
        if not t.get("name"):
            flag("tasks", i, "error", f"Tasks row {i + 1}: task name is required.")
        ck, pk = t.get("client", "").lower(), t.get("project", "").lower()
        if (ck, pk) not in project_keys:
            flag("tasks", i, "skip", f"Tasks row {i + 1}: project \"{t.get('project')}\" / client \"{t.get('client')}\" isn't in Projects — task skipped.")

    # Validate assignment emails
    for i, a in enumerate(assignments):
        em = a.get("user_email", "")
        if not em:
            flag("assignments", i, "error", f"Assignments row {i + 1}: user_email is required.")
        elif (await _user_by_email(db, em, tenant_id)) is None:
            flag("assignments", i, "skip", f"Assignments row {i + 1}: user \"{em}\" isn't a user here — assignment skipped.")

    return {
        "counts": {
            "clients": len(clients),
            "projects": len(projects),
            "tasks": len(tasks),
            "assignments": len(assignments),
        },
        "errors": errors,
        "warnings": warnings,
        "row_issues": row_issues,
        "data": data,  # echoed back so commit doesn't re-upload the file
    }


# ── Commit: create everything in dependency order ───────────────────────────
def _norm_status(value: str, allowed: set[str], default: str) -> str:
    v = _key(value)
    return v if v in allowed else default


async def commit_import(db: AsyncSession, data: dict[str, list[dict]], tenant_id: int) -> dict:
    """Create clients -> projects -> tasks -> assignments. Per-row failures are
    collected (the import doesn't abort on one bad row). Returns created counts +
    a per-row error list. Users are matched to existing accounts by email."""
    from app.crud.client import create_client
    from app.crud.project import create_project
    from app.crud.task import create_task, set_task_assignees
    from app.crud.project import set_project_roster, get_project_resource_ids
    from app.schemas import ClientCreate, ProjectCreate

    row_errors: list[str] = []
    created = {"clients": 0, "projects": 0, "tasks": 0, "assignments": 0}

    # 1) Clients (deduped by name; cache name->id).
    client_id_by_name: dict[str, int] = {}

    async def ensure_client(name: str, defaults: Optional[dict] = None) -> Optional[int]:
        if not name:
            return None
        key = name.lower()
        if key in client_id_by_name:
            return client_id_by_name[key]
        existing = await _client_by_name(db, name, tenant_id)
        if existing:
            client_id_by_name[key] = existing.id
            return existing.id
        d = defaults or {}
        cc = ClientCreate(
            name=name,
            client_type=_norm_status(d.get("type", ""), {"internal", "external"}, "external"),
            status=_norm_status(d.get("status", ""), {"active", "inactive", "suspended"}, "active"),
            company=d.get("company") or None,
            contact_name=d.get("contact_name") or None,
            contact_email=(d.get("contact_email") or None),
            contact_phone=d.get("contact_phone") or None,
            since=d.get("since") or None,
        )
        try:
            row = await create_client(db, cc, tenant_id)
            client_id_by_name[key] = row.id
            created["clients"] += 1
            return row.id
        except Exception as exc:  # noqa: BLE001
            row_errors.append(f"Client \"{name}\": {exc}")
            return None

    for c in data.get("clients", []):
        await ensure_client(c.get("name", ""), c)

    # 2) Projects (cache (client,name)->id; resolve PM emails to user ids).
    project_id_by_key: dict[tuple[str, str], int] = {}
    for p in data.get("projects", []):
        cname, pname = p.get("client", ""), p.get("name", "")
        if not pname or not cname:
            continue
        cid = await ensure_client(cname)
        if cid is None:
            row_errors.append(f"Project \"{pname}\": couldn't resolve client \"{cname}\".")
            continue
        manager_ids: list[int] = []
        for em in [e.strip() for e in p.get("managers", "").replace(",", ";").split(";") if e.strip()]:
            u = await _user_by_email(db, em, tenant_id)
            if u is not None:
                manager_ids.append(u.id)
        try:
            pc = ProjectCreate(
                name=pname,
                client_id=cid,
                billable_rate=(_to_decimal(p.get("billable_rate", "")) or Decimal("0")),
                code=p.get("code") or None,
                currency=p.get("currency") or None,
                budget_amount=_to_decimal(p.get("budget", "")),
                status=_norm_status(p.get("status", ""),
                                    {"planning", "in_progress", "on_hold", "completed"}, "planning"),
                manager_ids=manager_ids or None,
            )
            proj = await create_project(db, pc, tenant_id)
            project_id_by_key[(cname.lower(), pname.lower())] = proj.id
            created["projects"] += 1
        except Exception as exc:  # noqa: BLE001
            row_errors.append(f"Project \"{pname}\": {exc}")

    # 3) Tasks (cache (client,project,task)->id).
    task_id_by_key: dict[tuple[str, str, str], int] = {}
    for t in data.get("tasks", []):
        cname, pname, tname = t.get("client", ""), t.get("project", ""), t.get("name", "")
        if not tname:
            continue
        pid = project_id_by_key.get((cname.lower(), pname.lower()))
        if pid is None:
            row_errors.append(f"Task \"{tname}\": project \"{pname}\" (client \"{cname}\") not found — skipped.")
            continue
        try:
            task = await create_task(
                db, project_id=pid, tenant_id=tenant_id, name=tname,
                description=t.get("description") or None,
                priority=(_key(t.get("priority", "")) or None),
                status=(_norm_status(t.get("status", ""), {"to_do", "in_progress", "done"}, "to_do")),
            )
            task_id_by_key[(cname.lower(), pname.lower(), tname.lower())] = task.id
            created["tasks"] += 1
        except Exception as exc:  # noqa: BLE001
            row_errors.append(f"Task \"{tname}\": {exc}")

    # 4) Assignments. A blank task → grant the whole project (project roster);
    # a named task → add the user as that task's assignee. Users matched by email.
    for a in data.get("assignments", []):
        cname, pname, tname = a.get("client", ""), a.get("project", ""), a.get("task", "")
        email = a.get("user_email", "")
        u = await _user_by_email(db, email, tenant_id)
        if u is None:
            row_errors.append(f"Assignment: user \"{email}\" not found — skipped.")
            continue
        if tname:
            tid = task_id_by_key.get((cname.lower(), pname.lower(), tname.lower()))
            if tid is None:
                row_errors.append(f"Assignment: task \"{tname}\" not found — skipped.")
                continue
            try:
                from app.models.assignments import TaskAssignee
                task_obj = await db.get(Task, tid)
                # Merge this user into the task's current assignee set.
                cur_ids = set((await db.execute(
                    select(TaskAssignee.user_id).where(TaskAssignee.task_id == tid)
                )).scalars().all())
                cur_ids.add(u.id)
                await set_task_assignees(db, task_obj, list(cur_ids))
                created["assignments"] += 1
            except Exception as exc:  # noqa: BLE001
                row_errors.append(f"Assignment ({email} → task {tname}): {exc}")
        else:
            pid = project_id_by_key.get((cname.lower(), pname.lower()))
            if pid is None:
                row_errors.append(f"Assignment: project \"{pname}\" not found — skipped.")
                continue
            try:
                cur = set(await get_project_resource_ids(db, pid))
                cur.add(u.id)
                await set_project_roster(db, pid, sorted(cur))
                created["assignments"] += 1
            except Exception as exc:  # noqa: BLE001
                row_errors.append(f"Assignment ({email} → project {pname}): {exc}")

    return {"created": created, "errors": row_errors}
